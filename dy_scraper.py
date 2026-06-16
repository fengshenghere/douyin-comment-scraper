"""
抖音视频评论抓取核心引擎
通过浏览器CDP调用评论API（自动处理msToken + a_bogus签名）
"""

import sys, io
if sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from playwright.sync_api import sync_playwright
import json, time, re, sys
from urllib.parse import urlparse, parse_qs
import requests, os

CDP_URL = "http://127.0.0.1:28800"
COMMENT_API = "https://www.douyin.com/aweme/v1/web/comment/list/"
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36 Edg/149.0.0.0"
COOKIE_DOMAIN = ".douyin.com"


class DouyinScraper:
    def __init__(self, cdp_url=CDP_URL, cookie_str=None, cookie_file=None):
        self.cdp_url = cdp_url
        self.cookie_str = cookie_str
        self.cookie_file = cookie_file or os.path.join(os.path.dirname(__file__) or ".", "cookies.txt")
        self.playwright = None
        self.browser = None
        self.page = None

    @staticmethod
    def parse_cookies(cookie_str):
        """解析cookie字符串为Playwright格式"""
        cookies = []
        for pair in cookie_str.split("; "):
            if "=" in pair:
                k, v = pair.split("=", 1)
                cookies.append({"name": k, "value": v, "domain": COOKIE_DOMAIN, "path": "/"})
        return cookies

    def connect(self, timeout=15):
        """连接CDP浏览器并注入cookie（socket 超时保护）"""
        import socket
        orig_getaddrinfo = socket.getaddrinfo

        def _timeout_getaddrinfo(*args, **kwargs):
            old_timeout = socket.getdefaulttimeout()
            socket.setdefaulttimeout(timeout)
            try:
                return orig_getaddrinfo(*args, **kwargs)
            finally:
                socket.setdefaulttimeout(old_timeout)

        socket.getaddrinfo = _timeout_getaddrinfo
        try:
            self.playwright = sync_playwright().start()
            self.browser = self.playwright.chromium.connect_over_cdp(self.cdp_url)
        finally:
            socket.getaddrinfo = orig_getaddrinfo
        contexts = self.browser.contexts
        if not contexts:
            raise RuntimeError("没有浏览器上下文，请确认浏览器有已打开的标签页")
        self.page = contexts[0].pages[0] if contexts[0].pages else contexts[0].new_page()

        # CDP 模式：直接使用浏览器已有 cookie（无需注入）
        # 非 CDP 模式（launch）才需要注入 cookie
        if not hasattr(self, '_is_cdp_mode'):
            self._is_cdp_mode = True  # connect_over_cdp 就是 CDP 模式
        if self._is_cdp_mode:
            print(f"CDP 模式：复用浏览器已有 Cookie")
            return self

        # 以下是非 CDP 模式的 cookie 注入（保命兼容）
        cookie_str = self.cookie_str
        if not cookie_str and os.path.exists(self.cookie_file):
            try:
                with open(self.cookie_file, "r", encoding="utf-8") as f:
                    cookie_str = f.read().strip()
            except Exception:
                pass

        if cookie_str:
            cookies = self.parse_cookies(cookie_str)
            contexts[0].add_cookies(cookies)
            key_cookies = [c["name"] for c in cookies if c["name"] in ("sessionid", "passport_csrf_token", "sid_guard")]
            print(f"Cookies loaded: {len(cookies)} total, key={key_cookies}")

        return self

    def disconnect(self):
        if self.playwright:
            self.playwright.stop()
        self.page = None
        self.browser = None
        self.playwright = None

    # ---- 分享链接解析 ----
    @staticmethod
    def resolve_share_link(share_url):
        """短链 → https://www.douyin.com/video/{id}"""
        resp = requests.head(share_url, allow_redirects=True, timeout=15,
                             headers={"User-Agent": USER_AGENT})
        return resp.url

    @staticmethod
    def extract_video_id(url):
        """从视频页URL提取video_id"""
        # https://www.douyin.com/video/7650549224366558510?...
        m = re.search(r'/video/(\d+)', url)
        if m:
            return m.group(1)
        # 短链里的ID也行
        m = re.search(r'v\.douyin\.com/(\w+)/?', url)
        if m:
            return DouyinScraper.resolve_share_link(url)
        raise ValueError(f"无法从URL提取视频ID: {url}")

    # ---- 视频信息获取 ----
    @staticmethod
    def parse_share_text(text):
        """从分享文本中提取链接"""
        # 匹配 douyin.com 短链
        m = re.search(r'https?://v\.douyin\.com/[\w/-]+', text)
        if m:
            return m.group(0)
        # 匹配完整链接
        m = re.search(r'https?://www\.douyin\.com/video/\d+', text)
        if m:
            return m.group(0)
        raise ValueError("文本中未找到抖音分享链接")

    # ---- 评论抓取 ----
    @staticmethod
    def _safe_filename(title):
        """清理视频标题 → 安全文件名（≤10字，去特殊字符）"""
        import unicodedata
        # 去掉 " - 抖音" 后缀
        for suffix in (" - 抖音", "- 抖音", " - 抖音精选", "- 抖音精选"):
            if title.endswith(suffix):
                title = title[: -len(suffix)]
                break
        # 保留中文、字母、数字、空格
        clean = ""
        for ch in title:
            cat = unicodedata.category(ch)
            if cat.startswith("L") or cat.startswith("N") or ch in (" ", "_", "-"):
                clean += ch
            elif cat == "Zs":
                clean += " "
        # 合并多余空格，去首尾
        clean = " ".join(clean.split()).strip()
        if not clean:
            clean = "抖音视频"
        # 截断 ≤10 字符（按字符数，不是字节）
        if len(clean) > 10:
            clean = clean[:10]
        return clean.strip() or "抖音视频"

    def fetch_comments(self, video_id, max_pages=200, count=30):
        """
        从抖音视频获取评论列表，返回 (comments, video_title)
        通过浏览器 fetch 绕过 msToken / a_bogus 签名
        """
        # 先访问视频页面激活 cookie / token 上下文
        print(f"Loading video page: {video_id}")
        self.page.goto(f"https://www.douyin.com/video/{video_id}",
                       timeout=30000, wait_until="domcontentloaded")
        time.sleep(4)

        # 提取视频标题
        video_title = ""
        try:
            raw_title = self.page.title()
            video_title = self._safe_filename(raw_title)
            print(f"Video title: {raw_title[:50]}... → filename: {video_title}")
        except Exception:
            pass

        all_comments = []
        cursor = 0
        total_claimed = 0

        for page_num in range(1, max_pages + 1):
            result = self._fetch_page(video_id, cursor, count)
            status = result.get("status_code", -1)
            has_more = result.get("has_more", 0)
            next_cursor = result.get("cursor", 0)
            comments = result.get("comments", [])

            print(f"  Page {page_num}: cursor={cursor} has_more={has_more} "
                  f"next={next_cursor} got={len(comments)}")

            if status != 0 or not comments:
                break

            for c in comments:
                user_info = c.get("user", {})
                avatar_list = (user_info.get("avatar_thumb") or user_info.get("avatar_medium") or {}).get("url_list", [])
                all_comments.append({
                    "cid": c.get("cid"),
                    "text": c.get("text"),
                    "user": user_info.get("nickname"),
                    "uid": user_info.get("uid"),
                    "sec_uid": user_info.get("sec_uid"),
                    "avatar": avatar_list[0] if avatar_list else None,
                    "likes": c.get("digg_count", 0),
                    "time": c.get("create_time", 0),
                    "replies": c.get("reply_comment_total", 0),
                    "ip": c.get("ip_label"),
                })

            if total_claimed == 0 and comments:
                total_claimed = comments[0].get("item_comment_total",
                                                result.get("total", 0))

            if not has_more:
                break
            cursor = next_cursor
            time.sleep(0.2)

        print(f"\nDone: {len(all_comments)} comments (video claims {total_claimed})")
        return all_comments, video_title

    def _fetch_page(self, video_id, cursor, count):
        """单页API调用（在浏览器上下文内执行，自动签名）"""
        return self.page.evaluate(
            """
            async ([vid, cursor, count]) => {
                const base = 'https://www.douyin.com/aweme/v1/web/comment/list/';
                const params = new URLSearchParams({
                    device_platform: 'webapp',
                    aid: '6383',
                    channel: 'channel_pc_web',
                    aweme_id: vid,
                    cursor: String(cursor),
                    count: String(count),
                    item_type: '0',
                    update_version_code: '170400',
                    pc_client_type: '1',
                    version_code: '170400',
                    version_name: '17.4.0',
                    cookie_enabled: 'true',
                    screen_width: '2560',
                    screen_height: '1440',
                    browser_language: 'zh-CN',
                    browser_platform: 'Win32',
                    browser_name: 'Edge',
                    browser_version: '149.0.0.0',
                    browser_online: 'true',
                    platform: 'PC',
                    downlink: '10',
                    effective_type: '4g',
                    round_trip_time: '50'
                });
                const resp = await fetch(base + '?' + params.toString(), {
                    credentials: 'include'
                });
                if (!resp.ok) return { status_code: resp.status, has_more: 0, cursor: 0, comments: [] };
                try { return await resp.json(); } catch(e) { return { status_code: -1, has_more: 0, cursor: 0, comments: [] }; }
            }
            """,
            [video_id, cursor, count],
        )

    # ---- 子评论抓取 ----
    def fetch_sub_comments(self, video_id, comment_id, max_pages=10, count=10):
        """获取某条评论的子回复（通过浏览器API调用）"""
        all_replies = []
        cursor = 0

        for page_num in range(1, max_pages + 1):
            result = self._fetch_sub_page(video_id, comment_id, cursor, count)
            status = result.get("status_code", -1)
            has_more = result.get("has_more", 0)
            next_cursor = result.get("cursor", 0)
            comments = result.get("comments", [])

            if status != 0 or not comments:
                break

            for c in comments:
                user_info = c.get("user", {})
                avatar_list = (user_info.get("avatar_thumb") or {}).get("url_list", [])
                all_replies.append({
                    "cid": c.get("cid"),
                    "text": c.get("text"),
                    "user": user_info.get("nickname"),
                    "avatar": avatar_list[0] if avatar_list else None,
                    "likes": c.get("digg_count", 0),
                    "time": c.get("create_time", 0),
                    "ip": c.get("ip_label"),
                })

            if not has_more:
                break
            cursor = next_cursor
            time.sleep(0.15)

        return all_replies

    def _fetch_sub_page(self, video_id, comment_id, cursor, count):
        return self.page.evaluate(
            """
            async ([vid, cid, cursor, count]) => {
                const base = 'https://www.douyin.com/aweme/v1/web/comment/list/reply/';
                const params = new URLSearchParams({
                    device_platform: 'webapp',
                    aid: '6383',
                    channel: 'channel_pc_web',
                    aweme_id: vid,
                    comment_id: cid,
                    cursor: String(cursor),
                    count: String(count),
                    item_type: '0',
                    update_version_code: '170400',
                    pc_client_type: '1',
                    version_code: '170400',
                    version_name: '17.4.0',
                    cookie_enabled: 'true',
                    screen_width: '2560',
                    screen_height: '1440',
                    browser_language: 'zh-CN',
                    browser_platform: 'Win32',
                    browser_name: 'Edge',
                    take_black_list: '0',
                    is_blind: 'false',
                    is_business_comment: '0',
                    platform: 'PC'
                });
                const resp = await fetch(base + '?' + params.toString(), {
                    credentials: 'include'
                });
                if (!resp.ok) return { status_code: resp.status, has_more: 0, cursor: 0, comments: [] };
                try { return await resp.json(); } catch(e) { return { status_code: -1, has_more: 0, cursor: 0, comments: [] }; }
            }
            """,
            [video_id, comment_id, cursor, count],
        )

    # ---- 导出 ----
    def export_excel(self, comments, filepath):
        """导出评论到Excel（含样式）"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            print("请安装 openpyxl: pip install openpyxl")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "评论"

        # 表头
        headers = ["序号", "用户", "UID", "评论文本", "点赞", "回复数", "时间", "IP属地", "头像链接"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # 数据（按点赞降序）
        sorted_comments = sorted(comments, key=lambda x: x.get("likes", 0), reverse=True)

        from datetime import datetime
        for i, c in enumerate(sorted_comments, 1):
            ts = c.get("time", 0)
            dt_str = datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S") if ts else ""

            row_data = [
                i,
                c.get("user", ""),
                c.get("uid", ""),
                c.get("text", ""),
                c.get("likes", 0),
                c.get("replies", 0),
                dt_str,
                c.get("ip", ""),
                c.get("avatar", ""),
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i + 1, column=col, value=val)
                cell.border = thin_border
                if col in (1, 5, 6):
                    cell.alignment = Alignment(horizontal="center")
                if col == 4:
                    cell.alignment = Alignment(wrap_text=True)

        # 列宽
        widths = [6, 18, 18, 60, 10, 8, 20, 12, 50]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

        # 汇总行
        total_row = len(sorted_comments) + 2
        ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=f"共 {len(sorted_comments)} 条评论")
        ws.cell(row=total_row, column=5, value=sum(c.get("likes", 0) for c in sorted_comments))

        wb.save(filepath)
        print(f"Saved: {filepath}")
        return len(sorted_comments)


# ---- CLI ----
if __name__ == "__main__":
    import os

    # 输入
    if len(sys.argv) < 2:
        print("用法: python dy_scraper.py <抖音分享链接或分享文本> [输出文件名] [--subs] [--cookie <cookie字符串或文件路径>]")
        print("示例: python dy_scraper.py \"https://v.douyin.com/eZrw9DnN0aM/\"")
        print("      python dy_scraper.py \"分享文本含链接...\" result.xlsx --subs")
        print("      python dy_scraper.py \"链接\" --subs --cookie \"sessionid=xxx; ttwid=yyy\"")
        print("      python dy_scraper.py \"链接\" --subs --cookie cookies.txt")
        sys.exit(1)

    raw = sys.argv[1]
    out_file = None
    include_subs = "--subs" in sys.argv
    cookie_input = None

    # 解析参数
    i = 2
    while i < len(sys.argv):
        arg = sys.argv[i]
        if arg == "--cookie":
            if i + 1 < len(sys.argv):
                cookie_input = sys.argv[i + 1]
                i += 2
            else:
                i += 1
        elif arg == "--subs":
            i += 1
        elif not out_file and not arg.startswith("--"):
            out_file = arg
            i += 1
        else:
            i += 1

    # 解析输入
    if raw.startswith("http"):
        share_url = raw
    else:
        share_url = DouyinScraper.parse_share_text(raw)

    print(f"Share URL: {share_url}")

    # 处理cookie：如果是文件路径，读文件内容
    if cookie_input and not cookie_input.startswith("passport") and not ("=" in cookie_input and ";" in cookie_input):
        # 看起来是文件路径
        if os.path.exists(cookie_input):
            with open(cookie_input, "r", encoding="utf-8") as f:
                cookie_input = f.read().strip()
            print(f"Cookie loaded from file: {cookie_input}")

    scraper = DouyinScraper(cookie_str=cookie_input if cookie_input else None)
    try:
        scraper.connect()
        print(f"Connected to CDP: {CDP_URL}")

        # 解析短链
        full_url = DouyinScraper.resolve_share_link(share_url)
        video_id = DouyinScraper.extract_video_id(full_url)
        print(f"Video ID: {video_id}")
        print(f"Video URL: {full_url}")

        # 抓取评论
        comments, video_title = scraper.fetch_comments(video_id)

        # 子评论
        if include_subs and comments:
            print("\n=== Fetching sub-comments ===")
            total_subs = 0
            for c in comments:
                if c["replies"] > 0:
                    subs = scraper.fetch_sub_comments(video_id, c["cid"])
                    c["sub_comments"] = subs
                    total_subs += len(subs)
                    print(f"  {c['user']}: {len(subs)} sub-comments")
            print(f"Total sub-comments: {total_subs}")

        # 导出
        if not out_file:
            safe_name = f"{video_title}.xlsx" if video_title else f"dy_{video_id}.xlsx"
            out_file = os.path.join(os.path.dirname(__file__) or ".", safe_name)

        scraper.export_excel(comments, out_file)

        # 统计
        likes = sum(c["likes"] for c in comments)
        replied = sum(1 for c in comments if c["replies"] > 0)
        print(f"\nStats: {len(comments)} comments, {likes} likes, {replied} have replies")

    finally:
        scraper.disconnect()
